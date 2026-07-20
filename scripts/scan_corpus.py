#!/usr/bin/env python3
"""
scan_corpus.py — Scan a policy document folder and generate a mapping CSV for review.

Run this on your PC before ingestion. It will:
  1. Recurse into all subdirectories
  2. Extract PDF page counts and embedded titles where possible
  3. Auto-generate clean doc IDs and titles
  4. Detect overlaps with your existing 22 embedded docs
  5. Output a CSV you can review and edit before ingestion

Usage:
    python scan_corpus.py --docs-dir "C:/path/to/full_doc_database" --output corpus_mapping.xlsx

Requirements:
    pip install openpyxl PyPDF2
    (PyPDF2 is optional — script works without it, just skips page counts)
"""

import argparse
import csv
import re
import os
import sys
from pathlib import Path
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("openpyxl not installed — will output CSV instead of XLSX")
    print("Install with: pip install openpyxl")

try:
    from PyPDF2 import PdfReader
    HAS_PYPDF2 = True
except ImportError:
    HAS_PYPDF2 = False
    print("PyPDF2 not installed — will skip PDF page counts and metadata")
    print("Install with: pip install PyPDF2")


# ═══════════════════════════════════════════════════════════════════════════
# EXISTING DOCS — your 22 already-embedded documents
# ═══════════════════════════════════════════════════════════════════════════

EXISTING_DOCS = {
    "CAP24": "Climate Action Plan 2024",
    "CAP25": "Climate Action Plan 2025",
    "NBAP": "National Biodiversity Action Plan 2023-2030",
    "WAP24": "Water Action Plan 2024",
    "CLIMATE_ACT_2021": "Climate Action and Low Carbon Development (Amendment) Act 2021",
    "EU_NRL": "EU Nature Restoration Law (Regulation 2024/1991)",
    "EU_BIODIV_STRATEGY": "EU Biodiversity Strategy for 2030",
    "EU_WFD": "EU Water Framework Directive (2000/60/EC)",
    "CITIZENS_ASSEMBLY_BIODIV": "Report of the Citizens Assembly on Biodiversity Loss",
    "CAP_STRATEGIC_PLAN": "Ireland CAP Strategic Plan 2023-2027",
    "NAF_2024": "National Adaptation Framework 2024",
    "FOOD_VISION_2030": "Food Vision 2030",
    "FOREST_STRATEGY": "Ireland Forest Strategy 2023-2030",
    "NITRATES_AP5": "Fifth Nitrates Action Programme",
    "NECP": "Ireland National Energy and Climate Plan (NECP) 2021-2030",
    "EU_CLIMATE_LAW": "European Climate Law (Regulation 2021/1119)",
    "NPF": "National Planning Framework (Project Ireland 2040)",
    "HERITAGE_2030": "Heritage Ireland 2030",
    "CLEAN_AIR": "National Clean Air Strategy 2023",
    "CIRCULAR_ECONOMY": "Whole of Government Circular Economy Strategy",
    "EPA_SOE_2024": "EPA State of the Environment Report 2024",
    "FORESTRY_PROGRAMME": "Forestry Programme 2023-2027",
}

# Fingerprints for overlap detection (lowercase substrings → existing ID)
OVERLAP_FINGERPRINTS = {
    "climate-action-plan-2024": "CAP24",
    "climate_action_plan_2025": "CAP25",
    "climate action plan 2025": "CAP25",
    "4th_national_biodiversity_action_plan": "NBAP",
    "4th national biodiversity action plan": "NBAP",
    "water-action-plan-2024": "WAP24",
    "water action plan 2024": "WAP24",
    "nature restoration law": "EU_NRL",
    "eu biodiversity strategy for 2030": "EU_BIODIV_STRATEGY",
    "national-adaptation-framework-2024": "NAF_2024",
    "national adaptation framework 2024": "NAF_2024",
    "food-vision-2030": "FOOD_VISION_2030",
    "food vision 2030": "FOOD_VISION_2030",
    "irelands-forest-strategy-2023": "FOREST_STRATEGY",
    "irelands forest strategy 2023": "FOREST_STRATEGY",
    "fifth-nitrates-action-programme": "NITRATES_AP5",
    "fifth nitrates action programme": "NITRATES_AP5",
    "overview-of-irelands-fifth-nitrates": "NITRATES_AP5",
    "national-energy-and-climate-plan": "NECP",
    "national energy and climate plan": "NECP",
    "project-ireland-2040-npf": "NPF",
    "project ireland 2040": "NPF",
    "forest-strategy-implementation-plan-including-the-forestry-programme": "FORESTRY_PROGRAMME",
    "forestry programme 2023": "FORESTRY_PROGRAMME",
    "cap-strategic-plan-2023": "CAP_STRATEGIC_PLAN",
    "cap strategic plan 2023": "CAP_STRATEGIC_PLAN",
    "amendment_to_irelands_cap_strategic_plan": "CAP_STRATEGIC_PLAN",
    "heritage ireland 2030": "HERITAGE_2030",
    "clean air strategy": "CLEAN_AIR",
    "circular economy strategy": "CIRCULAR_ECONOMY",
    "state of the environment": "EPA_SOE_2024",
    "eu climate law": "EU_CLIMATE_LAW",
    "european climate law": "EU_CLIMATE_LAW",
    "climate action and low carbon development": "CLIMATE_ACT_2021",
    "citizens assembly": "CITIZENS_ASSEMBLY_BIODIV",
    "water framework directive": "EU_WFD",
}

# ═══════════════════════════════════════════════════════════════════════════
# DOMAIN CLASSIFICATION
# ═══════════════════════════════════════════════════════════════════════════

DOMAIN_KEYWORDS = {
    "climate": ["climate", "carbon", "emission", "energy", "just transition",
                "green deal", "adaptation framework", "low carbon", "renewable"],
    "biodiversity": ["biodiversity", "wildlife", "species", "habitat", "nature restoration",
                     "pollinator", "flora", "fauna", "birds", "hen harrier", "cites",
                     "aichi", "conservation", "peatland", "bog", "monuments", "cms",
                     "ramsar", "bern convention"],
    "water": ["water", "nitrate", "bathing", "groundwater", "drainage", "flood",
              "marine", "ocean", "maritime", "fisheries", "aquaculture", "port",
              "coastal", "foreshore", "seafood", "marpol", "ospar", "angling",
              "cframs", "harbour"],
    "agriculture": ["agricult", "food vision", "farm", "cap strategic", "ammonia",
                    "good agricultural", "ag climatise"],
    "forestry": ["forest", "coillte", "pearl mussel"],
    "cross_cutting": ["planning", "development", "landscape", "heritage", "rural",
                      "circular", "sdg", "sustainable", "environment", "epa",
                      "soil", "eia", "sea directive", "investment", "infrastructure",
                      "shared island"],
}

def classify_domain(text):
    tl = text.lower()
    for domain, keywords in DOMAIN_KEYWORDS.items():
        for kw in keywords:
            if kw in tl:
                return domain
    return "cross_cutting"


# ═══════════════════════════════════════════════════════════════════════════
# TITLE CLEANING
# ═══════════════════════════════════════════════════════════════════════════

def clean_title(filename):
    """Generate a clean human-readable title from a messy filename."""
    title = Path(filename).stem
    # Remove UUIDs
    title = re.sub(r'[-_]?[a-f0-9]{8}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{12}', '', title)
    # Remove trailing numbers/versions
    title = re.sub(r'[-_]+\d+$', '', title)
    title = re.sub(r'[-_]+v\d+$', '', title, flags=re.IGNORECASE)
    title = re.sub(r'[-_]+WEB$', '', title, flags=re.IGNORECASE)
    title = re.sub(r'[-_]+FINAL$', '', title, flags=re.IGNORECASE)
    title = re.sub(r'[-_]+updated[-_]+cover[-_]*\d*$', '', title, flags=re.IGNORECASE)
    # Replace hyphens/underscores with spaces
    title = title.replace('-', ' ').replace('_', ' ')
    # Clean up multiple spaces
    title = re.sub(r'\s+', ' ', title).strip()
    # Title case
    title = title.title()
    # Fix common casing issues
    for orig, repl in [("Eu ", "EU "), ("Ie ", "IE "), ("Uk ", "UK "),
                        ("Epa ", "EPA "), (" Of ", " of "), (" The ", " the "),
                        (" And ", " and "), (" For ", " for "), (" In ", " in "),
                        (" To ", " to "), (" A ", " a "), (" An ", " an "),
                        (" On ", " on "), ("Cbd", "CBD"), ("Cms", "CMS"),
                        ("Cites", "CITES"), ("Ospar", "OSPAR"),
                        ("Marpol", "MARPOL"), ("Nwrp", "NWRP"),
                        ("Cframs", "CFRAMS"), ("Sdg", "SDG"),
                        ("Eia", "EIA"), ("Dmap", "DMAP"),
                        ("Ore", "ORE"), ("Mara", "MARA"),
                        ("Npws", "NPWS"), ("Si No", "S.I. No"),
                        ("Sacs", "SACs"), ("Sea Statement", "SEA Statement"),
                        ("Ii ", "II "), ("Iii ", "III "), ("Cfp", "CFP"),
                        ("P3", "P3")]:
        title = title.replace(orig, repl)
    # Ensure first character is uppercase
    if title:
        title = title[0].upper() + title[1:]
    return title


def make_doc_id(title):
    """Generate a clean document ID from a title."""
    s = title.upper()
    s = re.sub(r'[^A-Z0-9\s]', '', s)
    words = s.split()
    skip = {'THE', 'OF', 'AND', 'FOR', 'IN', 'TO', 'A', 'AN', 'ON',
            'EUROPEAN', 'UNION', 'IRELAND', 'IRELANDS', 'IRISH', 'NATIONAL'}
    key_words = [w for w in words if w not in skip and len(w) > 1][:5]
    if not key_words:
        key_words = words[:3]
    doc_id = '_'.join(key_words)
    # Truncate
    return doc_id[:50]


# ═══════════════════════════════════════════════════════════════════════════
# PDF METADATA EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════

def extract_pdf_info(filepath):
    """Extract page count and embedded title from a PDF."""
    info = {"pages": None, "pdf_title": None}
    if not HAS_PYPDF2:
        return info
    try:
        reader = PdfReader(str(filepath))
        info["pages"] = len(reader.pages)
        meta = reader.metadata
        if meta and meta.title and len(meta.title.strip()) > 3:
            info["pdf_title"] = meta.title.strip()
    except Exception:
        pass
    return info


# ═══════════════════════════════════════════════════════════════════════════
# OVERLAP DETECTION
# ═══════════════════════════════════════════════════════════════════════════

def check_overlap(filename):
    """Check if a file overlaps with an existing embedded document."""
    fl = filename.lower()
    for fingerprint, existing_id in OVERLAP_FINGERPRINTS.items():
        if fingerprint in fl:
            return existing_id
    return None


# ═══════════════════════════════════════════════════════════════════════════
# SCANNER
# ═══════════════════════════════════════════════════════════════════════════

VALID_EXTENSIONS = {'.pdf', '.txt', '.html', '.htm', '.doc', '.docx'}
SKIP_EXTENSIONS = {'.xlsx', '.xls', '.csv', '.md', '.json', '.py', '.zip',
                   '.jpg', '.jpeg', '.png', '.gif', '.pptx'}

def scan_directory(docs_dir):
    """Recursively scan directory for policy documents."""
    docs_dir = Path(docs_dir)
    results = []

    for filepath in sorted(docs_dir.rglob('*')):
        if filepath.is_dir():
            continue

        ext = filepath.suffix.lower()
        if ext in SKIP_EXTENSIONS:
            continue
        if ext not in VALID_EXTENSIONS and ext not in {'.pdf'}:
            # Include unknown extensions but flag them
            pass

        # Relative path from root
        rel_path = filepath.relative_to(docs_dir)
        subfolder = str(rel_path.parent) if str(rel_path.parent) != '.' else ''

        # File size
        try:
            size_kb = filepath.stat().st_size / 1024
        except Exception:
            size_kb = 0

        # Check overlap
        overlap_id = check_overlap(str(filepath.name))

        # Clean title from filename
        title = clean_title(filepath.name)

        # Domain classification
        domain = classify_domain(filepath.name)

        # Generate doc ID
        doc_id = make_doc_id(title)

        # PDF metadata
        pdf_info = {"pages": None, "pdf_title": None}
        if ext == '.pdf':
            pdf_info = extract_pdf_info(filepath)

        # Use PDF embedded title if it's better than the filename-derived one
        best_title = title
        if pdf_info["pdf_title"] and len(pdf_info["pdf_title"]) > len(title) * 0.5:
            # Only use PDF title if it looks reasonable
            pt = pdf_info["pdf_title"]
            if not pt.startswith("Microsoft") and len(pt) < 200:
                best_title = pdf_info["pdf_title"]

        # Status
        if overlap_id:
            status = f"OVERLAP:{overlap_id}"
        elif ext in SKIP_EXTENSIONS or ext == '.htm':
            status = "SKIP"
        elif size_kb < 5:
            status = "SKIP:tiny"
        else:
            status = "NEW"

        results.append({
            "filename": filepath.name,
            "subfolder": subfolder,
            "full_path": str(filepath),
            "format": ext.replace('.', '').upper(),
            "size_kb": round(size_kb, 1),
            "pages": pdf_info["pages"],
            "status": status,
            "doc_id": doc_id if status == "NEW" else "",
            "clean_title": best_title if status == "NEW" else "",
            "domain": domain if status == "NEW" else "",
            "pdf_title": pdf_info["pdf_title"] or "",
            "notes": "",
        })

    return results


# ═══════════════════════════════════════════════════════════════════════════
# OUTPUT
# ═══════════════════════════════════════════════════════════════════════════

def write_xlsx(results, output_path):
    """Write results to a formatted XLSX for review."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Corpus Mapping"

    # Headers
    headers = ["Status", "Doc ID", "Clean Title", "Domain", "Filename",
               "Subfolder", "Format", "Size (KB)", "Pages", "PDF Title", "Notes"]
    header_fill = PatternFill('solid', fgColor='1A2D3D')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin_border = Border(
        bottom=Side(style='thin', color='243B4F')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Status colors
    status_colors = {
        "NEW": PatternFill('solid', fgColor='E8F5E9'),
        "OVERLAP": PatternFill('solid', fgColor='FFF3E0'),
        "SKIP": PatternFill('solid', fgColor='FFEBEE'),
    }
    domain_colors = {
        "climate": Font(color='3ABFBF'),
        "biodiversity": Font(color='4CAF6E'),
        "water": Font(color='5AA3D9'),
        "agriculture": Font(color='D4903A'),
        "forestry": Font(color='5B9E3A'),
        "cross_cutting": Font(color='8899AA'),
    }

    # Sort: NEW first, then OVERLAP, then SKIP
    sort_order = {"NEW": 0, "OVERLAP": 1, "SKIP": 2}
    results.sort(key=lambda r: (
        sort_order.get(r["status"].split(":")[0], 3),
        r["domain"],
        r["clean_title"],
    ))

    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r["status"])
        ws.cell(row=row_idx, column=2, value=r["doc_id"])
        ws.cell(row=row_idx, column=3, value=r["clean_title"])
        ws.cell(row=row_idx, column=4, value=r["domain"])
        ws.cell(row=row_idx, column=5, value=r["filename"])
        ws.cell(row=row_idx, column=6, value=r["subfolder"])
        ws.cell(row=row_idx, column=7, value=r["format"])
        ws.cell(row=row_idx, column=8, value=r["size_kb"])
        ws.cell(row=row_idx, column=9, value=r["pages"])
        ws.cell(row=row_idx, column=10, value=r["pdf_title"])
        ws.cell(row=row_idx, column=11, value=r["notes"])

        # Color-code status
        status_key = r["status"].split(":")[0]
        if status_key in status_colors:
            ws.cell(row=row_idx, column=1).fill = status_colors[status_key]

        # Color-code domain
        if r["domain"] in domain_colors:
            ws.cell(row=row_idx, column=4).font = domain_colors[r["domain"]]

        # Border and font
        for col in range(1, 12):
            ws.cell(row=row_idx, column=col).border = thin_border
            if col not in [1, 4]:
                ws.cell(row=row_idx, column=col).font = Font(size=9)

    # Column widths
    widths = [15, 35, 50, 15, 55, 30, 8, 10, 7, 50, 30]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f"A1:K{len(results) + 1}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2['A1'] = "Corpus Scan Summary"
    ws2['A1'].font = Font(bold=True, size=14)

    new_count = sum(1 for r in results if r["status"] == "NEW")
    overlap_count = sum(1 for r in results if r["status"].startswith("OVERLAP"))
    skip_count = sum(1 for r in results if r["status"].startswith("SKIP"))

    summary_data = [
        ("Total files scanned", len(results)),
        ("New documents (to ingest)", new_count),
        ("Overlaps with existing 22", overlap_count),
        ("Skipped files", skip_count),
        ("", ""),
        ("Domain Breakdown (new docs only)", ""),
    ]

    domain_counts = defaultdict(int)
    for r in results:
        if r["status"] == "NEW":
            domain_counts[r["domain"]] += 1
    for domain, count in sorted(domain_counts.items(), key=lambda x: -x[1]):
        summary_data.append((f"  {domain}", count))

    summary_data.extend([
        ("", ""),
        ("INSTRUCTIONS", ""),
        ("1. Review the 'Corpus Mapping' sheet", ""),
        ("2. Edit Doc ID and Clean Title as needed", ""),
        ("3. Change Status from NEW to SKIP for docs you don't want", ""),
        ("4. Change Domain if auto-classification is wrong", ""),
        ("5. Save and pass this file to the ingestion script", ""),
    ])

    for row_idx, (label, value) in enumerate(summary_data, 3):
        ws2.cell(row=row_idx, column=1, value=label)
        ws2.cell(row=row_idx, column=2, value=value)
        if label == "INSTRUCTIONS":
            ws2.cell(row=row_idx, column=1).font = Font(bold=True, size=12)

    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 15

    wb.save(output_path)
    print(f"\nSaved review spreadsheet → {output_path}")


def write_csv(results, output_path):
    """Fallback CSV output if openpyxl not available."""
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=[
            "status", "doc_id", "clean_title", "domain", "filename",
            "subfolder", "full_path", "format", "size_kb", "pages", "pdf_title", "notes"
        ])
        writer.writeheader()
        writer.writerows(results)
    print(f"\nSaved CSV → {output_path}")


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Scan policy document corpus")
    parser.add_argument("--docs-dir", required=True, help="Root folder containing policy documents")
    parser.add_argument("--output", default="corpus_mapping.xlsx", help="Output mapping file")
    args = parser.parse_args()

    docs_dir = Path(args.docs_dir)
    if not docs_dir.exists():
        print(f"Error: directory not found: {docs_dir}")
        sys.exit(1)

    print(f"Scanning: {docs_dir}")
    print(f"Recursing into subdirectories...")

    results = scan_directory(docs_dir)

    # Stats
    new_count = sum(1 for r in results if r["status"] == "NEW")
    overlap_count = sum(1 for r in results if r["status"].startswith("OVERLAP"))
    skip_count = sum(1 for r in results if r["status"].startswith("SKIP"))

    print(f"\nFound {len(results)} files:")
    print(f"  NEW:     {new_count}")
    print(f"  OVERLAP: {overlap_count}")
    print(f"  SKIP:    {skip_count}")

    # Output
    output_path = Path(args.output)
    if output_path.suffix == '.xlsx' and HAS_OPENPYXL:
        write_xlsx(results, str(output_path))
    else:
        csv_path = output_path.with_suffix('.csv')
        write_csv(results, str(csv_path))

    print(f"\nNext steps:")
    print(f"  1. Open {output_path.name} and review the mapping")
    print(f"  2. Edit Doc IDs and Clean Titles as needed")
    print(f"  3. Change Status to SKIP for any docs you don't want")
    print(f"  4. Save and use as input to the ingestion pipeline")


if __name__ == "__main__":
    main()